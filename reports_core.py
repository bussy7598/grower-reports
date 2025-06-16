import os
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import smtplib
from email.message import EmailMessage

def filter_master(path, start_date, end_date):
    df = pd.read_excel(path, header=1).dropna(how="all")
    df['Packed Date'] = pd.to_datetime(df['Packed Date'], dayfirst=True)
    
    mask = (df['Packed Date'] >= pd.to_datetime(start_date)) & \
        (df["Packed Date"] <= pd.to_datetime(end_date))\
        
    df = df.loc[mask].copy()

    df['GrowerName'] = (
        df['Supplier']
        .astype(str)
        .str.split("(", n=1)
        .str[0]
        .str.strip()
    )
    return df

def autosize_columns(ws, min_width=10, max_width=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        adjusted_width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width

def generate_reports(df, template_path, output_dir, growers=None, split_by_crop=False):
    if df is None or df.empty:
        return []

    os.makedirs(output_dir, exist_ok=True)
    paths = []

    expected_cols = [
        "Year Packed", "Packed Date", "Pack Week", "Crop", "Supplier",
        "TBC Ref. (Po No)", "Consignee", "Delivery Date", "Product", "Trays",
        "Net Weight", "Tray Price", "Total", "Grower Con Note (Or Load)",
        "Repacked", "Wasted", "Reconsigned", "MBM Kg Rate", "MBM Kg Charge ($)",
        "Commission %", "Commission Charge ($)", "Levy Charge ($) (ex Gst)",
        "Supermarket Charge %", "Supermarket Charge $", "Estimated Interstate $/Tray",
        "Estimated Interstate Charge ($)", "Estimated Fumigation $/Tray", "Estimated Fumigation Charge ($)",
        "Return To Farm Total ($)", "Net Return To Farm (Per Kg)"
    ]

    STYLE_MAP = {
        "A": {"number_format": None, "alignment": "center", "fill": None},
        "B": {"number_format": "dd/mm/yyyy", "alignment": "right", "fill": None},
        "C": {"number_format": "0", "alignment": "center", "fill": None},
        "D": {"number_format": None, "alignment": "left", "fill": None},
        "E": {"number_format": None, "alignment": "left", "fill": None},
        "F": {"number_format": None, "alignment": "left", "fill": None},
        "G": {"number_format": None, "alignment": "left", "fill": None},
        "H": {"number_format": "dd/mm/yyyy", "alignment": "right", "fill": None},
        "I": {"number_format": None, "alignment": "left", "fill": None},
        "J": {"number_format": "0", "alignment": "right", "fill": None},
        "K": {"number_format": "0.00", "alignment": "right", "fill": None},
        "L": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "M": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "N": {"number_format": None, "alignment": "left", "fill": None},
        "O": {"number_format": "0", "alignment": "right", "fill": None},
        "P": {"number_format": "0", "alignment": "right", "fill": None},
        "Q": {"number_format": "0", "alignment": "right", "fill": None},
        "R": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "S": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "T": {"number_format": "0.00", "alignment": "right", "fill": None},
        "U": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "V": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "W": {"number_format": "0.00", "alignment": "right", "fill": None},
        "X": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "Y": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "Z": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "AA": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "AB": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "AC": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
        "AD": {"number_format": "$#,##0.00", "alignment": "right", "fill": None},
    }

    alignments = {col: Alignment(horizontal=spec["alignment"]) for col, spec in STYLE_MAP.items()}

    SHEET_INDEX = 0
    START_ROW = 3
    PLACEHOLDERS = 2

    for grower, group in df.groupby("GrowerName"):
        if growers and grower not in growers:
            continue

        group = group.reindex(columns=expected_cols, fill_value="")
        group = group.sort_values("Packed Date")

        wb = load_workbook(template_path)
        ws = wb.worksheets[SHEET_INDEX]

        if split_by_crop:
            for crop, crop_group in group.groupby("Crop"):
                crop_group = crop_group.reindex(columns=expected_cols, fill_value="")
                crop_group = crop_group.sort_values("Packed Date")

                ws_crop = wb.copy_worksheet(ws)
                ws_crop.title = str(crop)[:31]

                ws_crop.delete_rows(START_ROW, PLACEHOLDERS)

                for r_off, row in enumerate(dataframe_to_rows(crop_group, index=False, header=True), START_ROW):
                    for c_off, val in enumerate(row, 1):
                        cell = ws_crop.cell(row=r_off, column=c_off, value=val)
                        col = cell.column_letter
                        nf = STYLE_MAP.get(col, {}).get("number_format")
                        if nf:
                            cell.number_format = nf
                        cell.alignment = alignments.get(col)
                        if STYLE_MAP.get(col, {}).get("fill"):
                            cell.fill = PatternFill(fill_type="solid", fgColor=STYLE_MAP[col]["fill"])

                autosize_columns(ws_crop)
            wb.remove(ws)

        else:
            ws.delete_rows(START_ROW, PLACEHOLDERS)

            for r_off, row in enumerate(dataframe_to_rows(group, False, False), START_ROW):
                for c_off, val in enumerate(row, 1):
                    cell = ws.cell(row=r_off, column=c_off, value=val)
                    col = cell.column_letter
                    nf = STYLE_MAP.get(col, {}).get("number_format")
                    if nf:
                        cell.number_format = nf
                    cell.alignment = alignments.get(col)
                    if STYLE_MAP.get(col, {}).get("fill"):
                        cell.fill = PatternFill(fill_type="solid", fgColor=STYLE_MAP[col]["fill"])

            autosize_columns(ws)

        out_path = os.path.join(output_dir, f"{grower} - TBC Grower Reports.xlsx")
        wb.save(out_path)
        paths.append(out_path)

    return paths

def send_reports(report_paths, email_map, smtp_cfg):

    today_str = datetime.today().strftime("%d.%m.%Y")
    with smtplib.SMTP(smtp_cfg["host"], smtp_cfg["port"]) as server:
        server.starttls()
        server.login(smtp_cfg["user"], smtp_cfg["password"])

        for path in report_paths:
            grower = os.path.basename(path).split(" - ")[0]
            to_addr = email_map.get(grower)
            if not to_addr:
                print(f"No email for {grower}, skipping.")
                continue

            msg = EmailMessage()
            msg["Subject"] =f"{grower} - TBC Grower Report {today_str}"
            msg["From"] = smtp_cfg["from_adress"]
            msg["To"] = to_addr
            msg.set_content(
                f"Hi {grower}, \n\n"
                "Please find attached your Return to Grower Report. \n"
                "Let us know if you have any questions. \n\n"
                "Regards, \n"
                "The Marketing Team"
            )

            with open(path, "rb") as f:
                data = f.read()
            msg.add_attachment(
                data,
                maintyp ="application",
                subtype="vnd.openxlmformats-officedocument.spreedsheetml.sheet",
                filename=os.path.basename(path)
            )
            server.send_message(msg)
            print(f"Sent report to {grower} at {to_addr}")